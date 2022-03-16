import datetime
import os
from typing import Optional

import django
import openpyxl
import pytz
from django.utils import timezone
from openpyxl.reader.workbook import Workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ships_tracker.settings.settings")
django.setup()

from ships_tracker.settings.settings import DATA_FILE_PATH
from ships_tracker.ships.models import Ship, ShipMovementHistory

timezone.now()


def upload_new_data():
    wookbook: Optional[Workbook] = None
    try:
        # commented code is for prod usage
        # today = datetime.datetime.utcnow()
        # wookbook = openpyxl.load_workbook(DATA_FILE_PATH + f"{today.year}{today.month}{today.day}.xlsx")
        wookbook = openpyxl.load_workbook(DATA_FILE_PATH + f"20190801.xlsx")  # test code
    except FileNotFoundError:
        print("No such file")

    worksheet = wookbook.active

    for i in range(1, worksheet.max_row):
        ship_data = [x[i].value for x in worksheet.iter_cols(1, worksheet.max_column)]

        code = ship_data[0]
        date_time = datetime.datetime.combine(
            ship_data[1], ship_data[2], tzinfo=pytz.UTC
        )
        width = ship_data[3]
        longitude = ship_data[4]
        name = ship_data[5]

        ship, was_created = Ship.objects.get_or_create(name=name, code=code)
        if was_created:
            ship.save()
        ShipMovementHistory(
            geo_datetime=date_time, longitude=longitude, width=width, ship=ship
        ).save()
